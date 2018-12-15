import openpyxl as op

'''
# Lecture note: Session 5 and 6

    Read and write data from excel files
    iteraton, iterable, and iterator
    List comprehension
'''

# ------------------------------ openpyxl ------------------------------#
# read_only = True : file is read much faster
# so for the huge file, read the file with read_only and write into a new file
# it will be much faster than writing into the same file

# book = op.load_workbook("sample_file.xlsx", read_only=True)
book = op.load_workbook("sample_file.xlsx")

# If you want to know the sheet name;
sheets = book.sheetnames
print (sheets) # it gives you all the lists of sheets

sheet = book.active
#sheet = book["Sheet1"]

v = sheet.cell(row=2, column=2).value
print (v)

# Write into a cell; column G, row 2
sheet["G2"] = 99

# Save the modified sheet in another file
book.save("new_file.xlsx")

'''
In-class exercise-1

    sample_file.xlsx contains grades for three teams A, B, and C. 
    Your program should do the following:
    
    Copy this file into your directory.
    Read data from the file into a list named “data”.
    Print data
    Calculate the sum of points obtained by each team and save the sum in another list.
    Write the total  for each team in column G.
    Save the workbook: book.save(“new_file.xlsx”)
    
    data ["A": [ grades ] , "B": [ grages ], "C": [ grades ]]
'''

maxRow = sheet.max_row # 4
maxCol = sheet.max_column # 6

data = []

for r in range(2, maxRow+1):

    team = sheet.cell(row=r, column=1).value
    grade = []
    sumgrade = 0

    for c in range(2, maxCol+1):

        g = sheet.cell(row=r, column=c).value
        grade.append(g)

        sumgrade += g

    data.append(team)
    data.append(grade)

    col_loc = "G"+str(r)
    sheet[col_loc] = sumgrade

print (data)

book.save("new_file2.xlsx")

# ------------------------------ parameter ordering ------------------------------#
# ------------------------------ keyword parameters ------------------------------#

def f(name, address):
    print ("Name:", name, "\n","Address: ", address)

f("Some name","101 Some Street, Town")

# keyword parameters
f(address="101 Some Street, Town", name="Some name")

# Default Parameter Values
def f2(name, address="Unknown"):
    print ("Name:", name, "\n","Address: ", address)

f2(name="Some name")

# ------------------------------ Breaking a loop ------------------------------#

num = None
sumNum = 0

while True:
    num = input()

    if num == "0": break

    sumNum += float(num)

print (sumNum)

# -------------------------- iteraton(range), iterable, and iterator --------------------------#
# -------------------------- List comprehension --------------------------#

x = range(4) # x = range(0,4)

#for i in x:
#    print (i)

y = iter([0,1,2,3])

for i in y:
    print (i) # It prints 0,1,2,3

for i in y:
    print (i) # It doesn't print anything

# This is difference bet. range and iterator

n = 6
x = []

for i in range(n+1):
    x.append(i)

# List comprehension; Another way of creating lists
# you can have nested for and if...
x = [y for y in range(n+1)]
print (x) # [0,1,2,3,4,5,6]

d = [(x,y) for x in range(3) for y in range(5)]
print (d) # [(0,0), (0,1), (0,2), (0,3), (0,4), (1,0), (1,1), (1,2),...(2,3), (2,4)]

s = [n for n in range(5)]
print (s) # [0,1,2,3,4]

s = [(x,y) for x in range(5) for y in range(5) if x*x+y*y>30]
print (s)