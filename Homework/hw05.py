import openpyxl as op
import numpy

def findSum(nums, target):

    # nums: a list of integers
    # target: target number

    num_seen = {}

    # Using hash-table; dictionary
    for i in range(len(nums)):

        complement = target - nums[i]

        if complement in num_seen:
            return [num_seen[complement], i]
        else:
            num_seen[nums[i]] = i

    return 0


def maxPrefix(strs):

    # find the longest common prefix string among an array of strings
    common = ""
    end_flag = False

    min_length = 99999

    for i in range(len(strs)):
        strs[i] = strs[i].replace(" ","")

        if len(strs[i]) < min_length:
            min_length = len(strs[i])

    idx = 0

    while idx < min_length:

        letter = strs[0][idx]

        for i in range(len(strs)):

            if (letter != strs[i][idx]):
                end_flag = True
                break

        if (end_flag == True):
            break
        else:
            common += letter
            idx += 1

    if (common == ""):
        return "NULL"
    else:
        return common

def getData(filename):

    # Load the file
    data = op.load_workbook(filename, read_only=True)
    sheet = data['Sheet1']

    # Initialization
    MV101 = []
    FIT101 = []
    LIT101 = []
    P101 = []
    FIT201 = []

    # check the length of series
    _len = 3 # starts from third row since the first and the second rows are not about the actual data
    while True:
        # Note that all series are of the same length
        if sheet.cell(row=_len, column=1).value is None:
            break
        _len += 1

    # check the column of each series
    col_idx = 1
    key = sheet.cell(row=2, column=col_idx).value

    col_FIT101 = -1
    col_LIT101 = -1
    col_MV101 = -1
    col_P101 = -1
    col_FIT201 = -1


    while key is not None:

        if key == 'FIT101':
            col_FIT101 = col_idx
        elif key == 'LIT101':
            col_LIT101 = col_idx
        elif key == 'MV101':
            col_MV101 = col_idx
        elif key == 'P101':
            col_P101 = col_idx
        elif key == 'FIT201':
            col_FIT201 = col_idx

        col_idx += 1
        key = sheet.cell(row=2, column=col_idx).value

    # Get data time series
    for i in range(3, _len):

        v = sheet.cell(row=i, column=col_FIT101).value
        FIT101.append(v)

        v = sheet.cell(row=i, column=col_LIT101).value
        LIT101.append(v)

        v = sheet.cell(row=i, column=col_MV101).value
        MV101.append(v)

        v = sheet.cell(row=i, column=col_P101).value
        P101.append(v)

        v = sheet.cell(row=i, column=col_FIT201).value
        FIT201.append(v)


    return [MV101,FIT101,LIT101,P101,FIT201]



def swatStates(MV101, FIT101, LIT101, P101, FIT201):

    checked_LIT101 = []
    checked_FIT101 = []
    checked_FIT201 = []

    len_of_data = len(MV101)

    for i in range(len_of_data):

        if P101[i] == 2 and MV101[i] == 2:
            checked_LIT101.append(LIT101[i])
            checked_FIT101.append(FIT101[i])
            checked_FIT201.append(FIT201[i])

    mean_checked_LIT101 = numpy.mean(checked_LIT101)
    mean_checked_FIT101 = numpy.mean(checked_FIT101)
    mean_checked_FIT201 = numpy.mean(checked_FIT201)

    var_checked_LIT101 = numpy.var(checked_LIT101)
    var_checked_FIT101 = numpy.var(checked_FIT101)
    var_checked_FIT201 = numpy.var(checked_FIT201)


    print("Device\tMean Variance")
    print("LIT101\t%.2f %.4f" % (mean_checked_LIT101, var_checked_LIT101))
    print("FIT101\t%.2f %.4f" % (mean_checked_FIT101, var_checked_FIT101))
    print("FIT201\t%.2f %.4f" % (mean_checked_FIT201, var_checked_FIT201))


def transient(data_series_MV101):

    trans_time = 0

    closing_begin = False
    closing_begin_time = 0

    for i in range(len(data_series_MV101)):

        if (data_series_MV101[i] == 0):

            if closing_begin == False:
                closing_begin = True
                closing_begin_time = i

            trans_time += 1

    return trans_time, closing_begin_time


if __name__ == "__main__":

    print ("Q1")

    nums = input()
    nums = nums.split(",")

    # type conversion
    for i in range(len(nums)):
        nums[i] = int(nums[i])

    target = input()
    target = int(target)

    output1 = findSum(nums=nums, target=target)
    print (output1)

    print ("Q2")

    strs = input()
    strs = strs.split(",")

    output2 = maxPrefix(strs=strs)

    print (output2)

    print ("Q3")

    dl = getData(filename="SWaTDatasetPartial.xlsx")

    swatStates(MV101=dl[0], FIT101=dl[1], LIT101=dl[2], P101=dl[3], FIT201=dl[4])

    # Compute the time to close the valve
    trans_time, closing_begin_time = transient(data_series_MV101=dl[0])

    if trans_time == 0:
        print("Could not find duration of transient state.")
    else:
        print("Closing begins at: %g" % closing_begin_time)
        print("Duration of transient state: %g seconds" % trans_time)