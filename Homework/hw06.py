from scipy.integrate import quad
from math import sin
import numpy as np
import openpyxl as op
from sklearn.cluster import KMeans
from scipy import linalg

#Q1
print("Q1")

# The limits a,b will be input from the keyboard
a, b = input().split(",")
a = float(a)
b = float(b)

ans, err = quad(lambda x: np.exp(-(x*x)) * sin(x), a=a, b=b, epsabs=1.0e-12, epsrel=1.0e-12)
ans = float(ans)

print("%0.6f" %(ans))

#Q2
print("Q2")

# Load the file
filename = "irisdata.xlsx"
data = op.load_workbook(filename, read_only=True)
sheet = data['Sheet1']

# Initialization
data = []

_len = 2
# starts from second row since the first row is not about the actual data
while True:
    # Note that all series are of the same length
    if sheet.cell(row=_len, column=1).value is None:
        break

    sepal_length = sheet.cell(row=_len, column=1).value
    sepal_width = sheet.cell(row=_len, column=2).value
    petal_length = sheet.cell(row=_len, column=3).value
    petal_width = sheet.cell(row=_len, column=4).value

    cur = [sepal_length, sepal_width, petal_length, petal_width]
    data.append(cur)

    _len += 1

kmeans = KMeans(n_clusters=3, n_init=100).fit(data)
sepal_lengths_of_the_centroid = [kmeans.cluster_centers_[0][0], kmeans.cluster_centers_[1][0], kmeans.cluster_centers_[2][0]]
sepal_lengths_of_the_centroid = sorted(sepal_lengths_of_the_centroid)

for i in range(0, 3):
    print("Centroid %d sepal length:%0.4f" %(i+1, sepal_lengths_of_the_centroid[i]))


#Q3
print("Q3")

n_row, n_col = input().split(",")
n_row = int(n_row)
n_col = int(n_col)

m_data = input().split(",")

data = []
one_row = []

# first change them to float
for i in range(len(m_data)):
    m_data[i] = float(m_data[i])

A = np.matrix(m_data)
A = np.reshape(A, (n_row, n_col))

inv_A = linalg.inv(A)

for i in range(len(inv_A)):
    for j in range(len(inv_A[i])):
        print("%0.2f" %(inv_A[i][j]))
